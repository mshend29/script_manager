from __future__ import annotations

from openpyxl import Workbook, load_workbook
import pytest

from core.project import Project
from core.project_settings import ProjectSettings
from import_engine.source_sync import SourceSyncEngine
from services.tracking_service import DELIVERED


def _write_source(path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SCRIPT"
    sheet.append([None, None, None, None, None])
    sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
    for row in rows:
        sheet.append(
            [
                row["in"],
                row["out"],
                row["dialogue"],
                row["character"],
                row["talent"],
            ]
        )
    workbook.save(path)
    workbook.close()


def _project(tmp_path) -> tuple[Project, object]:
    source = tmp_path / "source"
    source.mkdir()
    source_file = source / "AA23-第1集_中文.xlsx"
    project = Project(
        file_path=tmp_path / "identity-regression.smproj",
        settings=ProjectSettings(
            project_name="Identity Regression",
            source_folder=str(source),
            episode_before="第",
            episode_after="集",
        ),
        project_id="identity-regression",
    )
    project.save()
    return project, source_file


def _baseline_rows() -> list[dict[str, str]]:
    return [
        {
            "in": "00:00:01,000",
            "out": "00:00:02,000",
            "dialogue": "Halo",
            "character": "Hendra",
            "talent": "Brama",
        },
        {
            "in": "00:00:03,000",
            "out": "00:00:04,000",
            "dialogue": "Apa kabar?",
            "character": "Joko",
            "talent": "Dika",
        },
    ]


def _sync(project: Project, engine: SourceSyncEngine):
    report = engine.synchronize(project)
    assert not report.has_errors
    return report


def _active_dialogue_by_text(project: Project, text: str):
    with project.database.connect() as connection:
        return connection.execute(
            """
            SELECT d.id, d.dialog_uid, d.source_row, rs.is_recorded
            FROM dialogues AS d
            JOIN recording_status AS rs ON rs.dialogue_id = d.id
            WHERE d.is_active = 1
              AND d.dialog_text = ?
            """,
            (text,),
        ).fetchone()


def _mark_recorded(project: Project, dialogue_id: int) -> None:
    with project.database.connect() as connection:
        connection.execute(
            """
            UPDATE recording_status
            SET is_recorded = 1,
                recorded_at = '2026-09-02T10:00:00',
                updated_at = '2026-09-02T10:00:00'
            WHERE dialogue_id = ?
            """,
            (int(dialogue_id),),
        )


def _set_delivered(project: Project, dialogue_id: int) -> tuple[int, int, int]:
    with project.database.connect() as connection:
        scope = connection.execute(
            """
            SELECT d.episode_id, dc.character_id, dc.talent_id
            FROM dialogues AS d
            JOIN dialog_cast AS dc ON dc.dialogue_id = d.id
            WHERE d.id = ?
            LIMIT 1
            """,
            (int(dialogue_id),),
        ).fetchone()
        assert scope is not None
        assert scope["talent_id"] is not None
        values = (
            int(scope["episode_id"]),
            int(scope["talent_id"]),
            int(scope["character_id"]),
        )
        connection.execute(
            """
            INSERT INTO stem_status(
                episode_id, talent_id, character_id, status, note, updated_at
            )
            VALUES(?, ?, ?, ?, 'regression baseline', '2026-09-02T10:00:00')
            """,
            (*values, DELIVERED),
        )
        return values


def test_recorded_dialogue_text_revision_preserves_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    before = _active_dialogue_by_text(project, "Halo")
    assert before is not None
    _mark_recorded(project, int(before["id"]))

    revised = _baseline_rows()
    revised[0]["dialogue"] = "Halo semuanya"
    _write_source(source_file, revised)
    _sync(project, engine)

    after = _active_dialogue_by_text(project, "Halo semuanya")
    assert after is not None
    assert int(after["id"]) == int(before["id"])
    assert int(after["is_recorded"]) == 1


def test_recorded_dialogue_timecode_revision_preserves_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    before = _active_dialogue_by_text(project, "Halo")
    assert before is not None
    _mark_recorded(project, int(before["id"]))

    revised = _baseline_rows()
    revised[0]["in"] = "00:00:01,250"
    revised[0]["out"] = "00:00:02,250"
    _write_source(source_file, revised)
    _sync(project, engine)

    after = _active_dialogue_by_text(project, "Halo")
    assert after is not None
    assert int(after["id"]) == int(before["id"])
    assert int(after["is_recorded"]) == 1


def test_character_spelling_revision_preserves_dialogue_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    before = _active_dialogue_by_text(project, "Halo")
    assert before is not None
    _mark_recorded(project, int(before["id"]))

    revised = _baseline_rows()
    revised[0]["character"] = "Hendra Kepala Stasiun"
    _write_source(source_file, revised)
    _sync(project, engine)

    after = _active_dialogue_by_text(project, "Halo")
    assert after is not None
    assert int(after["id"]) == int(before["id"])
    assert int(after["is_recorded"]) == 1


def test_inserted_row_above_existing_dialogue_does_not_change_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    before = _active_dialogue_by_text(project, "Halo")
    assert before is not None
    _mark_recorded(project, int(before["id"]))

    inserted = [
        {
            "in": "00:00:00,100",
            "out": "00:00:00,900",
            "dialogue": "Dialog baru di atas",
            "character": "Teguh",
            "talent": "Vega",
        },
        *_baseline_rows(),
    ]
    _write_source(source_file, inserted)
    _sync(project, engine)

    after = _active_dialogue_by_text(project, "Halo")
    assert after is not None
    assert int(after["id"]) == int(before["id"])
    assert int(after["is_recorded"]) == 1
    assert int(after["source_row"]) == int(before["source_row"]) + 1


def test_moved_row_with_same_content_keeps_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    before = _active_dialogue_by_text(project, "Halo")
    assert before is not None
    _mark_recorded(project, int(before["id"]))

    moved = [rows[1], rows[0]]
    _write_source(source_file, moved)
    _sync(project, engine)

    after = _active_dialogue_by_text(project, "Halo")
    assert after is not None
    assert int(after["id"]) == int(before["id"])
    assert int(after["is_recorded"]) == 1
    assert int(after["source_row"]) != int(before["source_row"])


def test_duplicate_identical_source_rows_do_not_collide(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    duplicate = {
        "in": "00:00:01,000",
        "out": "00:00:02,000",
        "dialogue": "Halo",
        "character": "Hendra",
        "talent": "Brama",
    }
    _write_source(source_file, [duplicate, dict(duplicate)])

    report = _sync(project, engine)
    assert report.dialogues_added == 2

    with project.database.connect() as connection:
        rows = connection.execute(
            """
            SELECT id, dialog_uid
            FROM dialogues
            WHERE is_active = 1
            ORDER BY source_row, id
            """
        ).fetchall()

    assert len(rows) == 2
    assert int(rows[0]["id"]) != int(rows[1]["id"])
    assert str(rows[0]["dialog_uid"]) != str(rows[1]["dialog_uid"])


@pytest.mark.xfail(
    strict=True,
    reason="Current tracking invalidation is triggered by file fingerprint change.",
)
def test_formatting_only_source_change_does_not_reset_tracking(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    rows = _baseline_rows()
    _write_source(source_file, rows)
    _sync(project, engine)

    dialogue = _active_dialogue_by_text(project, "Halo")
    assert dialogue is not None
    episode_id, talent_id, character_id = _set_delivered(
        project,
        int(dialogue["id"]),
    )

    workbook = load_workbook(source_file)
    try:
        sheet = workbook["SCRIPT"]
        sheet.column_dimensions["C"].width = 48
        workbook.save(source_file)
    finally:
        workbook.close()

    prepared = engine.prepare(project)
    assert prepared.preview is not None
    assert prepared.preview.source_changed == 1
    assert prepared.preview.dialogues_added == 0
    assert prepared.preview.dialogues_removed == 0
    assert prepared.preview.text_changed == 0
    assert prepared.preview.cast_changed == 0

    engine.apply(project, prepared)

    with project.database.connect() as connection:
        status = connection.execute(
            """
            SELECT status
            FROM stem_status
            WHERE episode_id = ?
              AND talent_id = ?
              AND character_id = ?
            """,
            (episode_id, talent_id, character_id),
        ).fetchone()

    assert status is not None
    assert str(status["status"]) == DELIVERED
