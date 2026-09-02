from __future__ import annotations

import hashlib

from openpyxl import Workbook
import pytest

from core.project import Project
from core.project_settings import ProjectSettings
from import_engine.source_sync import SourceSyncEngine, SourceSyncError


def _write_source(path, *, dialogue: str = "Halo") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SCRIPT"
    sheet.append([None, None, None, None, None])
    sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
    sheet.append(
        [
            "00:00:01,000",
            "00:00:02,000",
            dialogue,
            "Hendra",
            "Brama",
        ]
    )
    workbook.save(path)
    workbook.close()


def _project(tmp_path):
    source = tmp_path / "source"
    source.mkdir()
    source_file = source / "AA23-第1集_中文.xlsx"
    _write_source(source_file)

    project = Project(
        file_path=tmp_path / "plan-test.smproj",
        settings=ProjectSettings(
            project_name="Plan Test",
            source_folder=str(source),
            episode_before="第",
            episode_after="集",
        ),
        project_id="plan-test",
    )
    project.save()
    return project, source_file


def _db_hash(project: Project) -> str:
    return hashlib.sha256(project.database_file.read_bytes()).hexdigest()


def test_prepare_plan_reuses_existing_dialogue_id_on_text_revision(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()

    initial = engine.synchronize(project)
    assert not initial.has_errors

    with project.database.connect() as connection:
        before = connection.execute(
            """
            SELECT id, dialog_uid, source_signature
            FROM dialogues
            WHERE is_active = 1
            """
        ).fetchone()
    assert before is not None

    _write_source(source_file, dialogue="Halo semuanya")
    prepared = engine.prepare(project)

    assert not prepared.has_errors
    assert prepared.plan is not None
    file_plan = prepared.plan.file_plans[str(source_file.resolve())]
    assert len(file_plan.matches) == 1
    assert not file_plan.additions
    assert not file_plan.removals
    assert file_plan.matches[0].existing.dialogue_id == int(before["id"])

    engine.apply(project, prepared)

    with project.database.connect() as connection:
        after = connection.execute(
            """
            SELECT id, dialog_uid, source_signature, dialog_text
            FROM dialogues
            WHERE is_active = 1
            """
        ).fetchone()

    assert int(after["id"]) == int(before["id"])
    assert str(after["dialog_uid"]) == str(before["dialog_uid"])
    assert str(after["source_signature"]) != str(before["source_signature"])
    assert str(after["dialog_text"]) == "Halo semuanya"


def test_apply_rejects_source_changed_after_preview(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    engine.synchronize(project)

    _write_source(source_file, dialogue="Revisi satu")
    prepared = engine.prepare(project)
    assert not prepared.has_errors
    assert prepared.plan is not None

    before_apply = _db_hash(project)
    _write_source(source_file, dialogue="Revisi dua")

    with pytest.raises(SourceSyncError, match="Source berubah setelah preview"):
        engine.apply(project, prepared)

    assert _db_hash(project) == before_apply
    assert prepared.backup_path == ""


def test_apply_rejects_database_changed_after_preview(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()
    engine.synchronize(project)

    _write_source(source_file, dialogue="Revisi")
    prepared = engine.prepare(project)
    assert not prepared.has_errors
    assert prepared.plan is not None

    with project.database.connect() as connection:
        connection.execute(
            """
            UPDATE dialogues
            SET dialog_text = 'Concurrent edit'
            WHERE is_active = 1
            """
        )

    before_apply = _db_hash(project)

    with pytest.raises(
        SourceSyncError,
        match="Database project berubah setelah preview",
    ):
        engine.apply(project, prepared)

    assert _db_hash(project) == before_apply
    assert prepared.backup_path == ""


def test_prepare_blocks_ambiguous_dialogue_lineage(tmp_path):
    project, source_file = _project(tmp_path)
    engine = SourceSyncEngine()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SCRIPT"
    sheet.append([None, None, None, None, None])
    sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
    sheet.append(["00:00:01,000", "00:00:02,000", "A", "Hendra", "Brama"])
    sheet.append(["00:00:03,000", "00:00:04,000", "B", "Joko", "Dika"])
    workbook.save(source_file)
    workbook.close()

    initial = engine.synchronize(project)
    assert not initial.has_errors

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SCRIPT"
    sheet.append([None, None, None, None, None])
    sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append(["00:00:11,000", "00:00:12,000", "A revised", "Alpha", "Brama"])
    sheet.append(["00:00:13,000", "00:00:14,000", "B revised", "Beta", "Dika"])
    workbook.save(source_file)
    workbook.close()

    prepared = engine.prepare(project)

    assert prepared.plan is not None
    assert prepared.plan.has_ambiguities
    assert prepared.has_errors
    assert any("lineage ambigu" in problem for problem in prepared.problems)
