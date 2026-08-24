from __future__ import annotations

import re
import zipfile
from pathlib import Path

from openpyxl import Workbook

from import_engine.inspector import WorkbookInspector
from import_engine.parser import ScriptParser


def _make_unsized_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Worksheet"
    sheet.append(["Episode 1"])
    sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
    sheet.append(
        [
            "00:00:01,000",
            "00:00:02,000",
            "Halo",
            "Hendra",
            "Brama",
        ]
    )
    workbook.save(path)

    rewritten = path.with_suffix(".rewritten.xlsx")
    with zipfile.ZipFile(path, "r") as source_zip:
        with zipfile.ZipFile(rewritten, "w") as target_zip:
            for info in source_zip.infolist():
                data = source_zip.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    data, replacements = re.subn(
                        rb"<dimension\b[^>]*/>",
                        b"",
                        data,
                        count=1,
                    )
                    assert replacements == 1
                target_zip.writestr(info, data)

    rewritten.replace(path)


def test_inspector_handles_read_only_workbook_without_dimension(tmp_path: Path) -> None:
    path = tmp_path / "AA23-第1集_中文.xlsx"
    _make_unsized_workbook(path)

    inspection = WorkbookInspector().inspect(path)

    assert inspection.sheet_names == ["Worksheet"]
    assert inspection.sheets[0].max_row >= 3
    assert inspection.sheets[0].max_column >= 5
    assert inspection.sheets[0].has_sample_data

    sample_values = [cell.value for cell in inspection.sheets[0].sample_cells]
    assert "DIALOG" in sample_values
    assert "Hendra" in sample_values


def test_parser_still_reads_unsized_workbook(tmp_path: Path) -> None:
    path = tmp_path / "AA23-第1集_中文.xlsx"
    _make_unsized_workbook(path)

    result = ScriptParser().parse(path, episode_number=1)

    assert result.layout.sheet_name == "Worksheet"
    assert result.layout.header_row == 2
    assert result.dialogue_count == 1
    assert result.rows[0].characters == ("Hendra",)
    assert result.rows[0].talents == ("Brama",)
    assert result.rows[0].status == "OK"
