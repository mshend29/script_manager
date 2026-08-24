from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from import_engine.parser import ScriptParser


class ScriptParserTests(unittest.TestCase):
    def _save_workbook(self, workbook: Workbook) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)

        path = Path(temp_dir.name) / "episode.xlsx"
        workbook.save(path)
        return path

    def test_detects_header_and_parses_multi_cast(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SCRIPT"

        sheet.append(["Catatan"])
        sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
        sheet.append(
            [
                "00:00:01,000",
                "00:00:02,000",
                "Kita pergi sekarang",
                "-Indah -Teguh",
                "Anisa - Brama",
            ]
        )
        sheet.append(
            [
                "00:00:03,000",
                "00:00:04,000",
                "Baik",
                "Bima",
                "Dika",
            ]
        )

        result = ScriptParser().parse(
            self._save_workbook(workbook),
            episode_number=7,
        )

        self.assertEqual(result.layout.detection, "header")
        self.assertEqual(result.dialogue_count, 2)

        first = result.rows[0]
        self.assertEqual(first.characters, ("Indah", "Teguh"))
        self.assertEqual(first.talents, ("Anisa", "Brama"))
        self.assertEqual(len(first.cast_pairs), 2)
        self.assertEqual(first.status, "OK")

    def test_legacy_a_to_e_fallback_requires_script_like_rows(self) -> None:
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(["Judul Episode"])
        sheet.append(["Mulai", "Selesai", "Teks Lokal", "Pemeran", "Pengisi"])
        sheet.append(
            [
                "00:00:01,000",
                "00:00:02,000",
                "Halo",
                "Bima",
                "Dika",
            ]
        )
        sheet.append(
            [
                "00:00:02,100",
                "00:00:03,000",
                "Apa kabar",
                "Bima",
                "Dika",
            ]
        )

        result = ScriptParser().parse(
            self._save_workbook(workbook),
            episode_number=8,
        )

        self.assertEqual(result.layout.detection, "legacy-a-e")
        self.assertEqual(result.dialogue_count, 2)

    def test_cast_count_mismatch_is_not_guessed(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
        sheet.append(
            [
                "00:00:01,000",
                "00:00:02,000",
                "Bersama",
                "Hendra - Joko",
                "Brama",
            ]
        )

        result = ScriptParser().parse(
            self._save_workbook(workbook),
            episode_number=9,
        )

        row = result.rows[0]
        self.assertIn("CAST_COUNT_MISMATCH", row.status)
        self.assertEqual(row.cast_pairs, ())
        self.assertTrue(result.warnings)

    def test_real_source_cast_patterns_are_preserved_safely(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SCRIPT"
        sheet.append(["IN", "OUT", "DIALOG", "TOKOH", "TALENT"])
        sheet.append(
            [
                "00:00:27,633",
                "00:00:30,100",
                "-kunci ini ke mereka\n-Betul",
                "-Bima, \n-Indah",
                "-Dika\n-Anisa",
            ]
        )
        sheet.append(
            [
                "00:01:12,867",
                "00:01:14,100",
                "Ngapain capek-capek begini",
                "ibu baju hitam &\nibu baju biru",
                "Fitri &\nAnisa",
            ]
        )
        sheet.append(
            [
                "00:00:58,200",
                "00:00:59,933",
                "-Mau apa\n-Tenang saja",
                "- Satria\n- Bapak sepatu bot",
                "- Vega\n- Vega",
            ]
        )
        sheet.append(
            [
                "00:00:10,",
                "00:00:11,",
                "[Reaksi]",
                "[Crowded Anak kecil]",
                "Anisa, Fitri, Anggraini",
            ]
        )
        sheet.append(
            [
                "00:00:27,",
                "00:00:29,",
                "[Reaksi]",
                "[Crowded]",
                "All",
            ]
        )

        result = ScriptParser().parse(
            self._save_workbook(workbook),
            episode_number=45,
        )

        self.assertEqual(result.dialogue_count, 5)

        first = result.rows[0]
        self.assertEqual(first.characters, ("Bima", "Indah"))
        self.assertEqual(first.talents, ("Dika", "Anisa"))
        self.assertEqual(
            [(pair.character, pair.talent) for pair in first.cast_pairs],
            [("Bima", "Dika"), ("Indah", "Anisa")],
        )

        ampersand = result.rows[1]
        self.assertEqual(
            ampersand.characters,
            ("ibu baju hitam", "ibu baju biru"),
        )
        self.assertEqual(ampersand.talents, ("Fitri", "Anisa"))

        repeated = result.rows[2]
        self.assertEqual(repeated.talents, ("Vega", "Vega"))
        self.assertEqual(
            [(pair.character, pair.talent) for pair in repeated.cast_pairs],
            [("Satria", "Vega"), ("Bapak sepatu bot", "Vega")],
        )
        self.assertNotIn("CAST_COUNT_MISMATCH", repeated.status)

        crowd = result.rows[3]
        self.assertEqual(crowd.characters, ("Crowded Anak kecil",))
        self.assertEqual(crowd.talents, ("Anisa", "Fitri", "Anggraini"))
        self.assertEqual(len(crowd.cast_pairs), 3)
        self.assertIn("MULTI_TALENT", crowd.status)
        self.assertEqual(crowd.time_in, "00:00:10,000")
        self.assertEqual(crowd.time_out, "00:00:11,000")

        generic = result.rows[4]
        self.assertEqual(generic.characters, ("Crowded",))
        self.assertEqual(generic.talents, ())
        self.assertEqual(len(generic.cast_pairs), 1)
        self.assertEqual(generic.cast_pairs[0].talent, "")
        self.assertIn("GENERIC_TALENT", generic.status)
        self.assertIn("MISSING_TALENT", generic.status)


if __name__ == "__main__":
    unittest.main()
