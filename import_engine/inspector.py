from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class WorkbookInspectionError(RuntimeError):
    pass


@dataclass(frozen=True)
class SampleCell:
    sheet_name: str
    row: int
    column: int
    coordinate: str
    value: Any


@dataclass
class SheetInspection:
    name: str
    state: str
    max_row: int
    max_column: int
    sample_cells: list[SampleCell] = field(default_factory=list)

    @property
    def has_sample_data(self) -> bool:
        return bool(self.sample_cells)


@dataclass
class WorkbookInspection:
    file_path: str
    file_name: str
    sheets: list[SheetInspection] = field(default_factory=list)

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets]

    @property
    def visible_sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets if sheet.state == "visible"]

    @property
    def total_sample_cells(self) -> int:
        return sum(len(sheet.sample_cells) for sheet in self.sheets)


class WorkbookInspector:
    """Read workbook structure without making parser assumptions."""

    def __init__(
        self,
        *,
        sample_row_limit: int = 80,
        sample_column_limit: int = 40,
        sample_cell_limit: int = 250,
    ) -> None:
        self.sample_row_limit = sample_row_limit
        self.sample_column_limit = sample_column_limit
        self.sample_cell_limit = sample_cell_limit

    def inspect(self, file_path: str | Path) -> WorkbookInspection:
        path = Path(file_path).expanduser()

        if not path.exists():
            raise WorkbookInspectionError(f"File tidak ditemukan: {path}")

        if not path.is_file():
            raise WorkbookInspectionError(f"Path bukan file: {path}")

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:  # noqa: BLE001
            raise WorkbookInspectionError(
                f"Workbook tidak dapat dibuka: {exc}"
            ) from exc

        try:
            sheets = [self._inspect_sheet(sheet) for sheet in workbook.worksheets]

            if not sheets:
                raise WorkbookInspectionError("Workbook tidak memiliki worksheet.")

            return WorkbookInspection(
                file_path=str(path.resolve()),
                file_name=path.name,
                sheets=sheets,
            )
        finally:
            workbook.close()

    def _inspect_sheet(self, sheet) -> SheetInspection:
        max_row, max_column = self._resolve_sheet_dimensions(sheet)

        row_limit = min(max_row, self.sample_row_limit) if max_row else 0
        column_limit = (
            min(max_column, self.sample_column_limit) if max_column else 0
        )

        sample_cells: list[SampleCell] = []

        if row_limit and column_limit:
            for row_index, row_values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=row_limit,
                    min_col=1,
                    max_col=column_limit,
                    values_only=True,
                ),
                start=1,
            ):
                for column_index, value in enumerate(row_values, start=1):
                    if value is None or value == "":
                        continue

                    sample_cells.append(
                        SampleCell(
                            sheet_name=sheet.title,
                            row=row_index,
                            column=column_index,
                            coordinate=(
                                f"{get_column_letter(column_index)}{row_index}"
                            ),
                            value=self._serializable_value(value),
                        )
                    )

                    if len(sample_cells) >= self.sample_cell_limit:
                        break

                if len(sample_cells) >= self.sample_cell_limit:
                    break

        return SheetInspection(
            name=sheet.title,
            state=str(sheet.sheet_state),
            max_row=max_row,
            max_column=max_column,
            sample_cells=sample_cells,
        )

    @staticmethod
    def _resolve_sheet_dimensions(sheet) -> tuple[int, int]:
        """Return worksheet bounds, including read-only sheets without <dimension>.

        Some client-generated Excel files omit the worksheet ``dimension`` element.
        ``openpyxl`` then exposes ``max_row`` / ``max_column`` as ``None`` in
        read-only mode even though cell data is present. Force a one-time dimension
        calculation so inspection does not incorrectly report an empty worksheet.
        """
        max_row = int(sheet.max_row or 0)
        max_column = int(sheet.max_column or 0)

        if max_row and max_column:
            return max_row, max_column

        try:
            sheet.calculate_dimension(force=True)
        except (AttributeError, TypeError, ValueError):
            # Keep the inspector non-destructive. If a third-party workbook cannot
            # calculate dimensions, the parser can still produce the actionable
            # validation error later in the pipeline.
            return max_row, max_column

        return int(sheet.max_row or 0), int(sheet.max_column or 0)

    @staticmethod
    def _serializable_value(value: Any) -> Any:
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()

        if isinstance(value, (str, int, float, bool)) or value is None:
            return value

        return str(value)
