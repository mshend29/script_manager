from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from import_engine.normalizer import (
    clean_text,
    looks_like_timecode,
    normalize_dialogue_key,
    normalize_key,
    normalize_timecode,
    split_cast_value,
)


class ScriptParseError(RuntimeError):
    pass


@dataclass(frozen=True)
class ColumnLayout:
    sheet_name: str
    header_row: int
    start_row: int
    in_column: int
    out_column: int
    dialogue_column: int
    character_column: int
    talent_column: int
    detection: str = "header"


@dataclass(frozen=True)
class CastPair:
    character: str
    talent: str


@dataclass(frozen=True)
class ParsedDialogueRow:
    source_row: int
    time_in: str
    time_out: str
    dialogue: str
    raw_character: str
    raw_talent: str
    characters: tuple[str, ...]
    talents: tuple[str, ...]
    cast_pairs: tuple[CastPair, ...]
    dialog_uid: str
    status: str

    @property
    def is_multi_character(self) -> bool:
        return len(self.characters) > 1


@dataclass
class ScriptParseResult:
    file_path: str
    file_name: str
    episode_number: int
    layout: ColumnLayout
    rows: list[ParsedDialogueRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def dialogue_count(self) -> int:
        return len(self.rows)


_HEADER_ALIASES: dict[str, set[str]] = {
    "in": {
        "in",
        "time in",
        "timein",
        "start",
        "start time",
        "tc in",
        "timecode in",
    },
    "out": {
        "out",
        "time out",
        "timeout",
        "end",
        "end time",
        "tc out",
        "timecode out",
    },
    "dialogue": {
        "dialog",
        "dialogue",
        "dialog text",
        "subtitle",
        "text",
        "terjemahan",
        "translation",
    },
    "character": {
        "tokoh",
        "character",
        "characters",
        "karakter",
        "role",
        "roles",
    },
    "talent": {
        "talent",
        "voice talent",
        "voice",
        "va",
        "dubber",
        "pengisi suara",
    },
}


class ScriptParser:
    """
    Parse source workbooks into neutral dialogue rows.

    This stage deliberately does not resolve character/talent against the database.
    It preserves raw values, normalizes names, identifies safe positional cast pairs,
    and generates a stable dialog_uid for later synchronization.
    """

    def __init__(
        self,
        *,
        header_scan_rows: int = 25,
        header_scan_columns: int = 20,
    ) -> None:
        self.header_scan_rows = header_scan_rows
        self.header_scan_columns = header_scan_columns

    def parse(
        self,
        file_path: str | Path,
        *,
        episode_number: int,
    ) -> ScriptParseResult:
        path = Path(file_path).expanduser()

        if not path.exists():
            raise ScriptParseError(f"File tidak ditemukan: {path}")

        if not path.is_file():
            raise ScriptParseError(f"Path bukan file: {path}")

        try:
            episode = int(episode_number)
        except (TypeError, ValueError) as exc:
            raise ScriptParseError(
                f"Nomor episode tidak valid: {episode_number}"
            ) from exc

        if episode < 1:
            raise ScriptParseError("Nomor episode harus lebih besar dari 0.")

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:  # noqa: BLE001
            raise ScriptParseError(f"Workbook tidak dapat dibuka: {exc}") from exc

        try:
            if not workbook.worksheets:
                raise ScriptParseError("Workbook tidak memiliki worksheet.")

            layout = self._detect_layout(workbook)
            worksheet = workbook[layout.sheet_name]

            result = ScriptParseResult(
                file_path=str(path.resolve()),
                file_name=path.name,
                episode_number=episode,
                layout=layout,
            )

            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=layout.start_row,
                    values_only=True,
                ),
                start=layout.start_row,
            ):
                parsed = self._parse_row(
                    values,
                    row_number=row_number,
                    episode_number=episode,
                    layout=layout,
                    warnings=result.warnings,
                )

                if parsed is not None:
                    result.rows.append(parsed)

            if not result.rows:
                raise ScriptParseError(
                    f"Tidak ada dialog yang dapat dibaca pada sheet "
                    f"'{layout.sheet_name}'."
                )

            return result
        finally:
            workbook.close()

    # =========================================================
    # LAYOUT DETECTION
    # =========================================================

    def _detect_layout(self, workbook) -> ColumnLayout:
        sheets = sorted(
            workbook.worksheets,
            key=lambda sheet: 0 if sheet.sheet_state == "visible" else 1,
        )

        candidates: list[tuple[int, ColumnLayout]] = []

        for sheet in sheets:
            header_candidate = self._detect_header_layout(sheet)
            if header_candidate is not None:
                score, layout = header_candidate
                candidates.append((score, layout))

        if candidates:
            candidates.sort(key=lambda item: item[0], reverse=True)
            return candidates[0][1]

        # Safe compatibility fallback for the historical source format:
        # row 2 header, row 3+ data, A-E = IN/OUT/DIALOG/TOKOH/TALENT.
        # The fallback is accepted only if actual row values look like script data.
        for sheet in sheets:
            if self._looks_like_legacy_a_to_e(sheet):
                return ColumnLayout(
                    sheet_name=sheet.title,
                    header_row=2,
                    start_row=3,
                    in_column=1,
                    out_column=2,
                    dialogue_column=3,
                    character_column=4,
                    talent_column=5,
                    detection="legacy-a-e",
                )

        raise ScriptParseError(
            "Struktur naskah tidak dikenali. Header IN/OUT/DIALOG/TOKOH/TALENT "
            "tidak ditemukan dan fallback A-E tidak cocok."
        )

    def _detect_header_layout(
        self,
        sheet,
    ) -> tuple[int, ColumnLayout] | None:
        best: tuple[int, dict[str, int], int] | None = None

        for row_index, row in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=self.header_scan_rows,
                min_col=1,
                max_col=self.header_scan_columns,
                values_only=True,
            ),
            start=1,
        ):
            mapping: dict[str, int] = {}

            for column_index, value in enumerate(row, start=1):
                token = self._header_token(value)
                if not token:
                    continue

                for field_name, aliases in _HEADER_ALIASES.items():
                    if field_name in mapping:
                        continue

                    if token in aliases:
                        mapping[field_name] = column_index
                        break

            if "dialogue" not in mapping:
                continue

            score = len(mapping) * 10

            if "in" in mapping and "out" in mapping:
                score += 8

            if "character" in mapping:
                score += 5

            if "talent" in mapping:
                score += 3

            if best is None or score > best[0]:
                best = (score, mapping, row_index)

        if best is None:
            return None

        score, mapping, header_row = best

        # Require enough structure to avoid treating an arbitrary TEXT column
        # as a script header.
        if score < 35:
            return None

        return (
            score,
            ColumnLayout(
                sheet_name=sheet.title,
                header_row=header_row,
                start_row=header_row + 1,
                in_column=mapping.get("in", 0),
                out_column=mapping.get("out", 0),
                dialogue_column=mapping["dialogue"],
                character_column=mapping.get("character", 0),
                talent_column=mapping.get("talent", 0),
                detection="header",
            ),
        )

    @staticmethod
    def _header_token(value: object) -> str:
        token = normalize_key(value)
        token = token.replace("_", " ")
        return " ".join(token.split())

    @staticmethod
    def _looks_like_legacy_a_to_e(sheet) -> bool:
        matched_rows = 0
        inspected_rows = 0

        for row in sheet.iter_rows(
            min_row=3,
            max_row=15,
            min_col=1,
            max_col=5,
            values_only=True,
        ):
            inspected_rows += 1

            time_in = row[0] if len(row) > 0 else None
            time_out = row[1] if len(row) > 1 else None
            dialogue = clean_text(row[2] if len(row) > 2 else None)

            if dialogue and (
                looks_like_timecode(time_in)
                or looks_like_timecode(time_out)
            ):
                matched_rows += 1

        return inspected_rows > 0 and matched_rows >= 2

    # =========================================================
    # ROW PARSING
    # =========================================================

    def _parse_row(
        self,
        values: tuple[Any, ...],
        *,
        row_number: int,
        episode_number: int,
        layout: ColumnLayout,
        warnings: list[str],
    ) -> ParsedDialogueRow | None:
        raw_in = self._value_at(values, layout.in_column)
        raw_out = self._value_at(values, layout.out_column)
        raw_dialogue = self._value_at(values, layout.dialogue_column)
        raw_character_value = self._value_at(values, layout.character_column)
        raw_talent_value = self._value_at(values, layout.talent_column)

        if all(
            value in (None, "")
            for value in (
                raw_in,
                raw_out,
                raw_dialogue,
                raw_character_value,
                raw_talent_value,
            )
        ):
            return None

        dialogue = clean_text(raw_dialogue)

        if not dialogue:
            warnings.append(f"Row {row_number}: DIALOG kosong, row dilewati.")
            return None

        time_in = normalize_timecode(raw_in)
        time_out = normalize_timecode(raw_out)
        raw_character = clean_text(raw_character_value)
        raw_talent = clean_text(raw_talent_value)

        characters = tuple(split_cast_value(raw_character))
        talents = tuple(split_cast_value(raw_talent))
        cast_pairs = self._safe_cast_pairs(characters, talents)

        statuses: list[str] = []

        if not time_in:
            statuses.append("MISSING_IN")

        if not time_out:
            statuses.append("MISSING_OUT")

        if not characters:
            statuses.append("MISSING_CHARACTER")

        if characters and not talents:
            statuses.append("MISSING_TALENT")

        if characters and talents and len(characters) != len(talents):
            statuses.append("CAST_COUNT_MISMATCH")
            warnings.append(
                f"Row {row_number}: jumlah TOKOH ({len(characters)}) dan "
                f"TALENT ({len(talents)}) berbeda."
            )

        dialog_uid = build_dialog_uid(
            episode_number=episode_number,
            characters=characters,
            talents=talents,
            time_in=time_in,
            time_out=time_out,
            dialogue=dialogue,
        )

        return ParsedDialogueRow(
            source_row=row_number,
            time_in=time_in,
            time_out=time_out,
            dialogue=dialogue,
            raw_character=raw_character,
            raw_talent=raw_talent,
            characters=characters,
            talents=talents,
            cast_pairs=cast_pairs,
            dialog_uid=dialog_uid,
            status=" | ".join(statuses) if statuses else "OK",
        )

    @staticmethod
    def _safe_cast_pairs(
        characters: tuple[str, ...],
        talents: tuple[str, ...],
    ) -> tuple[CastPair, ...]:
        if not characters:
            return ()

        if not talents:
            return tuple(CastPair(character=name, talent="") for name in characters)

        if len(characters) != len(talents):
            # Do not guess a many-to-many relationship when counts differ.
            return ()

        return tuple(
            CastPair(character=character, talent=talent)
            for character, talent in zip(characters, talents, strict=True)
        )

    @staticmethod
    def _value_at(values: tuple[Any, ...], column: int) -> Any:
        if column < 1:
            return None

        index = column - 1

        if index >= len(values):
            return None

        return values[index]


def build_dialog_uid(
    *,
    episode_number: int,
    characters: tuple[str, ...] | list[str],
    talents: tuple[str, ...] | list[str],
    time_in: str,
    time_out: str,
    dialogue: str,
) -> str:
    """Build stable source identity without using Excel row numbers."""
    character_keys = sorted(
        {normalize_key(name) for name in characters if normalize_key(name)}
    )

    if character_keys:
        cast_key = "+".join(character_keys)
    else:
        # Character-less rows are unusual. Talent is used only as a fallback
        # so two different unresolved lines do not collapse onto one identity.
        talent_keys = sorted(
            {normalize_key(name) for name in talents if normalize_key(name)}
        )
        cast_key = "talent:" + "+".join(talent_keys) if talent_keys else "unknown"

    payload = "|".join(
        (
            str(int(episode_number)),
            cast_key,
            normalize_timecode(time_in),
            normalize_timecode(time_out),
            normalize_dialogue_key(dialogue),
        )
    )

    return hashlib.sha1(payload.encode("utf-8")).hexdigest()  # noqa: S324
